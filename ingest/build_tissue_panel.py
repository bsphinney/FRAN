"""build_tissue_panel.py — load the Yue et al. 2026 tissue-enriched protein panel into the corpus.

SOURCE. Yue, Jiang, Li, Luo et al., "Spatial distribution of the proteome in the human body and in
cancers", Nature 656:227 (2026), doi 10.1038/s41586-026-10660-y, open access. Supplementary Table 5
sheet A: 1,717 tissue-enriched proteins scored across 64 tissue types.

WHY THIS SOURCE. It is DIA-MS, the same modality as this corpus, so marker DETECTABILITY transfers.
The Human Protein Atlas alternative is antibody-based and semi-quantitative; the paper itself notes
that limitation. A marker we cannot detect is not a marker.

Measured against FRAN before committing to it (2026-08-12):
  * 1,688 of 1,717 marker genes (98.3%) appear somewhere in delimp_proteins
  * 44 of 57 tissues retain >=5 observable markers
  * 5,507 human acquisitions available to label

A protein is assigned to the tissue where its z-score is highest. The runner-up z is kept because the
margin between best and second-best is what separates a specific marker from a merely-elevated one.

TWO MATCHING TRAPS, both handled here rather than in the scorer:
  * `delimp_proteins.gene` is ';'-separated for shared protein groups -- 1,239,649 rows contain a
    ';'. An exact `gene = 'CYP1A2'` match silently misses every one of those.
  * Capitalisation is inconsistent ('Hsd17b13' alongside 'HSD17B13') because the corpus spans
    species and engines. Matching is therefore case-insensitive, which is safe because scoring is
    restricted to human runs.

    python ingest/build_tissue_panel.py <supp_table_5.xlsx>          # dry run
    python ingest/build_tissue_panel.py <supp_table_5.xlsx> --apply
"""
import argparse
import functools
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
print = functools.partial(print, flush=True)   # noqa: A001

SOURCE = "Yue2026_Nature_656_227_SuppTable5A"

DDL = """
CREATE TABLE IF NOT EXISTS delimp_tissue_marker_panel (
    tissue        text NOT NULL,
    gene          text NOT NULL,
    gene_upper    text NOT NULL,   -- match key; the corpus mixes 'HSD17B13' and 'Hsd17b13'
    uniprot       text,
    z_score       double precision,
    runner_up_z   double precision,
    source        text NOT NULL,
    PRIMARY KEY (source, tissue, gene)
);
CREATE INDEX IF NOT EXISTS idx_tmp_gene   ON delimp_tissue_marker_panel (gene_upper);
CREATE INDEX IF NOT EXISTS idx_tmp_tissue ON delimp_tissue_marker_panel (tissue);
"""


def _conn():
    import psycopg2
    from refresh_leaderboards import _token
    return psycopg2.connect(
        host=os.environ.get("DELIMP_PG_HOST", "pgfarm.library.ucdavis.edu"), port=5432,
        dbname=os.environ.get("DELIMP_PG_DB", "uc-davis-genome-center-proteomics-core/delimp"),
        user=os.environ.get("DELIMP_PG_USER", "genome-proteomics-service-account"),
        password=_token(), sslmode="require", connect_timeout=30,
        options="-c statement_timeout=600000")


def parse(xlsx):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "tissue enriched" in s.lower()), None)
    if sheet is None:
        sys.exit(f"no 'tissue enriched' sheet in {xlsx}; found {wb.sheetnames}")
    rows = list(wb[sheet].iter_rows(values_only=True))
    tissues = [str(h).strip() for h in rows[0][2:] if h]

    def num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    out = []
    for r in rows[1:]:
        uniprot, gene = r[0], r[1]
        if not gene:
            continue
        vals = [(tissues[i], num(v)) for i, v in enumerate(r[2:2 + len(tissues)])]
        vals = [(t, z) for t, z in vals if z is not None]
        if not vals:
            continue
        vals.sort(key=lambda x: -x[1])
        best_t, best_z = vals[0]
        runner = vals[1][1] if len(vals) > 1 else None
        g = str(gene).strip()
        out.append((best_t, g, g.upper(), str(uniprot).strip() if uniprot else None,
                    best_z, runner, SOURCE))
    return out, tissues


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--apply", action="store_true")
    a = ap.parse_args()
    if not os.path.exists(a.xlsx):
        sys.exit(f"no such file: {a.xlsx}")

    rows, tissues = parse(a.xlsx)
    by_t = {}
    for t, *_ in rows:
        by_t[t] = by_t.get(t, 0) + 1
    print(f"parsed {len(rows):,} markers over {len(by_t)} tissues "
          f"(sheet declares {len(tissues)} tissue columns)")
    for t, n in sorted(by_t.items(), key=lambda x: -x[1])[:8]:
        print(f"   {t:26s} {n:>4}")

    conn = _conn(); conn.autocommit = False
    cur = conn.cursor()
    cur.execute("SET LOCAL lock_timeout = '10s'")

    if not a.apply:
        print("\nDRY RUN — re-run with --apply.")
        conn.rollback(); conn.close(); return

    cur.execute(DDL); conn.commit()
    cur.execute("DELETE FROM delimp_tissue_marker_panel WHERE source = %s", (SOURCE,))
    import psycopg2.extras
    psycopg2.extras.execute_values(cur, """
        INSERT INTO delimp_tissue_marker_panel
          (tissue, gene, gene_upper, uniprot, z_score, runner_up_z, source)
        VALUES %s ON CONFLICT (source, tissue, gene) DO NOTHING""", rows, page_size=500)
    conn.commit()
    print(f"stored {len(rows):,} markers")

    # Observability is NOT checked here. It needs a gene-level join over delimp_proteins, the
    # service account cannot CREATE TEMP TABLE on this database, and the scorer has to do that join
    # anyway -- so it reports coverage as a by-product instead of duplicating the work.
    print("panel stored; run ingest/predict_tissue.py for observability + scoring")

    import versions as V
    V.record_run(cur, "tissue_marker_panel", "1.0.0",
                 notes=f"{len(rows)} markers, {len(by_t)} tissues")
    conn.commit(); conn.close()
    print("DONE")


if __name__ == "__main__":
    main()
